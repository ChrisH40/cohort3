import invoice_sheet_functions
import openpyxl
from openpyxl.styles import Font
import datetime


def invoice_number():
    while True:
        try:
            invoice_id = int(input("Please enter your invoice ID number: "))
        except ValueError:
            print("Not a valid number. Please re-enter.")
            continue
        else:
            return invoice_id


def invoice_generator(file):
    sheets = invoice_sheet_functions.sheet_reader(file)
    while True:
        try:
            invoice_ID = invoice_number()
            invoice = invoice_sheet_functions.invoice_to_dict(invoice_ID, sheets)
        except:
            print("Invoice not found.")
        else:
            return invoice_printer(invoice)


def invoice_printer(inv):
    invoice = openpyxl.Workbook()
    sheet = invoice.active
    sheet.title = f"Invoice{inv['invoice']['INV_ID']}"

    def apt_number(inv):
        if inv['customer']['APT_NUM'] == None:
            return ""
        else:
            return f"#{inv['customer']['APT_NUM']},"
    
    title_style = Font(sz=22.0, b=True, color="1dcc4b")
    header_style = Font(sz=16.0, b=True)
    footer_style = Font(sz=16.0, b=True, i=True)
    bolded_style = Font(b=True)
    info_style = Font(sz=10.0)

    title = sheet['B2']
    title.value = "Billy's Bespin Bazaar"
    title.font = title_style

    invoice_date = sheet['L2']
    invoice_date.value = f"{inv['invoice']['INV_DATE'].strftime('%d %B %Y')}"
    invoice_date.font = info_style
    invoice_num = sheet['L3']
    invoice_num.value = f"#{inv['invoice']['INV_ID']}"
    invoice_num.font = info_style

    cust_name = sheet['B5']
    cust_name.value = f"{inv['customer']['F_NAME']} {inv['customer']['L_NAME']}"
    cust_name.font = header_style
    cust_address1 = sheet['B6']
    cust_address1.value = f"{apt_number(inv)} {inv['customer']['ST_NUM']} {inv['customer']['ST_NAME']}"
    cust_address1.font = header_style
    cust_address2 = sheet['B7']
    cust_address2.value = f"{inv['customer']['CITY']} {inv['customer']['PROV']}, {inv['customer']['CTRY']}"
    cust_address2.font = header_style
    cust_address3 = sheet['B8']
    cust_address3.value = f"{inv['customer']['P_CODE']}"
    cust_address3.font = header_style

    header_prod = sheet['B11']
    header_prod.value = "Product"
    header_prod.font = bolded_style
    header_desc = sheet['E11']
    header_desc.value = "Description"
    header_desc.font = bolded_style
    header_unit = sheet['I11']
    header_unit.value = "Unit Cost"
    header_unit.font = bolded_style
    header_qty = sheet['K11']
    header_qty.value = "Qty"
    header_qty.font = bolded_style
    header_total = sheet['M11']
    header_total.value = "Line Total"
    header_total.font = bolded_style

    def prod_printer(inv):
        column_start = 2
        row_counter = 13
        inv_total = 0
        for item in inv['products']:
            sheet.cell(row = row_counter, column = column_start).value = item['product']
            sheet.cell(row = row_counter, column = column_start + 3).value = item['desc']
            sheet.cell(row = row_counter, column = column_start + 7).value = f"${item['cost']}"
            sheet.cell(row = row_counter, column = column_start + 9).value = item['quantity']
            line_total = round((item['quantity'] * item['cost']), 2)
            sheet.cell(row = row_counter, column = column_start + 11).value = f"${line_total}"
            inv_total += line_total
            row_counter += 2
        sheet.cell(row = row_counter + 2, column = 12).value = "Total:"
        sheet.cell(row = row_counter + 2, column = 13).font = bolded_style
        sheet.cell(row = row_counter + 2, column = 13).value = f"${round(inv_total, 2)}"

    prod_printer(inv)

    footer = sheet['C31']
    footer.value = "May the force be with you..."
    footer.font = footer_style

    invoice.save(
        f'/code/cohort3/src/python/comp-230/{sheet.title.lower()}.xlsx')


invoice_generator('/code/cohort3/src/python/comp-230/invoice_data.xlsx')
