from openpyxl import load_workbook


def sheet_reader(file):
    workbook = load_workbook(file)
    worksheets = workbook.sheetnames
    sheets = dict()
    for worksheet in worksheets:
        read_worksheet = workbook[worksheet]
        sheets[worksheet] = sheet_to_dict(read_worksheet)
    return sheets


def sheet_to_dict(sheet):
    headers = sheet[1]
    sheet_dict = dict()
    for row in sheet:
        if row != headers:
            row_dict = dict()
            index = 0
            for cell in row:
                if headers[index].value != None:
                    row_dict[headers[index].value] = cell.value
                    index += 1
            sheet_dict[row[0].value] = row_dict
    return sheet_dict


def invoice_to_dict(inv, sheets):
    customers = sheets['Customers']
    invoices = sheets['Invoices']
    products = sheets['Products']
    line_items = sheets['Line Items']
    invoice = dict()
    invoice['invoice'] = invoices[inv]
    invoice['customer'] = customers[invoices[inv]['CUST_ID']]
    prod_list = []
    for value in line_items.values():
        if inv == value['INV_ID']:
            prod_dict = dict()
            prod_id = value['PROD_ID']
            prod_dict['product'] = products[prod_id]['PROD_NAME']
            prod_dict['desc'] = products[prod_id]['PROD_DESC']
            prod_dict['cost'] = products[prod_id]['UNIT_PRICE']
            prod_dict['quantity'] = value['QTY']
            prod_list.append(prod_dict)
            invoice['products'] = prod_list
    return invoice
